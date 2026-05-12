from PySide6.QtPrintSupport import QPrinter
from PySide6.QtGui import QTextDocument, QFont
from openpyxl import load_workbook 
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os
from copy import copy

class Exportador:
    @staticmethod
    def generar_html(datos):
        with open("templates/reporte.html", "r", encoding="utf-8") as f:
            html = f.read()
            
        fecha = datetime.now().strftime("%d/%m/%Y")
        numero = Exportador.generar_nro_reporte()
        resumen = datos["resumen"]
        
        filas_html = ""
        for fila in datos["tabla"]:
            filas_html += "<tr>"
            for valor in fila:
                filas_html += f"<td>{valor}</td>"
            filas_html += "</tr>"

        # reemplazar los datos
        html = html.replace("{{empresa}}", datos["empresa"])
        html = html.replace("{{fecha}}", fecha)
        html = html.replace("{{numero}}", numero)
        html = html.replace("{{saldo_inicial}}", resumen["saldo_inicial"])
        html = html.replace("{{ingresos}}", resumen["ingresos"])
        html = html.replace("{{egresos}}", resumen["egresos"])
        html = html.replace("{{saldo_total}}", resumen["saldo_total"])
        html = html.replace("{{observaciones}}", datos["observaciones"])
        html = html.replace("{{filas}}", filas_html)
        
        return html
    
    @staticmethod
    def exportar_pdf(html, ruta):
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(ruta)
        
        doc = QTextDocument()
        doc.setDefaultFont(QFont("Arial", 12))
        doc.setHtml(html)

        doc.print_(printer)
    
    @staticmethod
    def generar_excel(datos, ruta):
        wb = load_workbook("templates/reporte_base.xlsx")
        ws = wb.active
        ws.title = "Reporte"
        
        fecha = datetime.now().strftime("%d/%m/%Y")
        numero = Exportador.generar_nro_reporte()
        resumen = datos["resumen"]
        
        # 🔹 llenar datos fijos
        ws["C3"] = "REPORTE CONTROL DE CAJA"
        ws["C6"] = f"Empresa: {datos['empresa']}"
        ws["G3"] = fecha    # fecha de reporte
        ws["G4"] = numero   # nro de reporte
        
        # RESUMEN 
        ws["G6"] = Exportador.conversor_nro(resumen["saldo_inicial"])# lo demas se calcula
        
        ws["C12"] = datos["observaciones"]

        # 🔹 tabla
        fila_inicio = 15  # debajo del header  
        fila_total_original = 18  
        fila_marco_original = 19   
        
        filas_datos = len(datos["tabla"])
        # fila_total_nueva = fila_inicio + filas_datos

        fila_final = fila_inicio + filas_datos - 1
        fila_total_nueva = fila_final + 1
        fila_marco_nueva = fila_total_nueva + 1
        
        # mover totales
        ws.move_range(
            f"A{fila_total_original}:G{fila_total_original}",
            rows=fila_total_nueva - fila_total_original,
            cols=0
        )
        
        # mover marco inferior
        ws.move_range(
            f"A{fila_marco_original}:G{fila_marco_original}",
            rows=fila_marco_nueva - fila_marco_original,
            cols=0
        )
        
        # llenar tabla
        for i, fila in enumerate(datos["tabla"], start=fila_inicio):
            
            fila_modelo = 15 if i % 2 == 0 else 16 # tu fila bonita 
            
            # copiar estilo de celda
            Exportador.copiar_estilo(ws,fila_modelo, i)
            
            ws[f"C{i}"] = fila[0]  # Fecha
            ws[f"D{i}"] = fila[1]  # Concepto
            ws[f"E{i}"] = Exportador.conversor_nro(fila[2])  # Ingreso
            ws[f"F{i}"] = Exportador.conversor_nro(fila[3])  # Egreso
            
            if i == fila_inicio:
                ws[f"G{i}"] = f"=G6+E{i}-F{i}"
            else:
                ws[f"G{i}"] = f"=G{i-1}+E{i}-F{i}"
                
            # if not any(fila):
            #     continue
             
        # bordes laterales
        for i in range(fila_inicio, fila_final + 1):

            # borde izquierdo
            ws.cell(row=i, column=2).border = copy(
                ws.cell(row=fila_inicio, column=2).border
            )

            # borde derecho
            ws.cell(row=i, column=8).border = copy(
                ws.cell(row=fila_inicio, column=8).border
            )
        
        # formula actualizada final
        fila_final = fila_inicio + filas_datos - 1
        
        ws[f"E{fila_total_nueva}"] = f"=SUM(E{fila_inicio}:E{fila_final})"
        ws[f"F{fila_total_nueva}"] = f"=SUM(F{fila_inicio}:F{fila_final})"
        
        # resumen de tabla formulas
        ws["G7"] = f"=E{fila_total_nueva}"  # total ingresos
        ws["G8"] = f"=F{fila_total_nueva}"  # total egresos
        
        # combina celdas de total al moverlas
        ws.merge_cells(f"C{fila_total_nueva}:D{fila_total_nueva}")
        # centramos el texto
        ws[f"C{fila_total_nueva}"].alignment = Alignment(horizontal="center")
        # borde grueso arriba
        borde = Border(top=Side(style="thick"))

        for col in ["C", "D", "E", "F", "G"]:
            # ws[f"{col}{fila_total_nueva}"].border = borde
            celda = ws[f"{col}{fila_total_nueva}"]

            celda.border = Border(
                left=copy(celda.border.left),
                right=copy(celda.border.right),
                bottom=copy(celda.border.bottom),
                top=Side(style="thick")
            )
            
        wb.save(ruta)
        
    @staticmethod
    def copiar_estilo(ws, fila_origen, fila_destino):
        for col in range(1, 8): # columnas A G
            origen = ws.cell(row=fila_origen, column=col)
            destino = ws.cell(row=fila_destino,column=col)
            
            if origen.has_style:
                destino.font = copy(origen.font)
                destino.border = copy(origen.border)
                destino.fill = copy(origen.fill)
                destino.number_format = copy(origen.number_format)
                destino.alignment = copy(origen.alignment)
            
    def conversor_nro(valor):
        if not valor:
            return 0
        
        print("VALOR ORIGINAL:", valor)
        
        try:
            # limpiamos texto
            valor = str(valor)
            valor = valor.replace("S/", "")
            valor = valor.replace(",", "")
            valor = valor.strip()
            
            return float(valor)
        except:
            return 0
   
    @staticmethod
    def generar_nro_reporte():
        ruta = "datos/reportes_contador.txt"
        
        if not os.path.exists(ruta):
            with open(ruta, "w") as f:
                f.write("1")
            return "0001"
        
        with open(ruta, "r+") as f:
            numero = int(f.read())
            f.seek(0)
            f.write(str(numero + 1))
            
        return str(numero).zfill(4)